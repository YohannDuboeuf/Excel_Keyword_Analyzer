from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor, as_completed
from oletools.olevba import VBA_Parser
from openpyxl import load_workbook

import os
import shutil
import argparse
import gc
import multiprocessing
import subprocess
import xlrd

import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")

# === folder by default ===
EXCEL_FOLDER = './assets/excel'
MACRO_FOLDER = './assets/macro'
RESULT_FOLDER = './assets/macro_trouves'


def process_all_excels_in_parallel(keyword: str, source_folder: str, output_folder: str):

    """
    Processes all Excel files in the source folder.
    Extracts macros and analyzes formulas to search for a specific keyword.
    """
    target_folder = os.path.join(output_folder, keyword.lower())

    # clean output folder before processing
    shutil.rmtree(target_folder, ignore_errors=True)
    os.makedirs(target_folder, exist_ok=True)

    os.makedirs(MACRO_FOLDER, exist_ok=True)

    # Search for .xls and .xlsm files
    all_excel_files = []
    for root, _, files in os.walk(source_folder):
        for file in files:
            path = os.path.join(root, file)
            if file.lower().endswith('.xlsm'):
                all_excel_files.append(path)
            elif file.lower().endswith('.xls'):
                converted = convert_xls_to_xlsx(path, source_folder)
                if converted:
                    all_excel_files.append(converted)

    # Parallel processing
    with ThreadPoolExecutor(max_workers=os.cpu_count() or 4) as executor:
        futures = [executor.submit(process_excel_file, path, keyword, target_folder) for path in all_excel_files]
        for _ in tqdm(as_completed(futures), total=len(futures), desc="Processing Excel files"):
            pass

    print(f"\n✅ Analysis complete. Results saved in : {target_folder}")


def analyze_formulas_xls(file_path: str, keyword: str, target_folder: str) -> bool:
    """
    Analyzes formulas in a .xls file to search for the keyword.
    """
    copied = False
    normalized_keyword = keyword.lower().lstrip("=")
    try:
        wb = xlrd.open_workbook(file_path)
        for sheet in tqdm(wb.sheets(), desc=f"[{os.path.basename(file_path)}] Sheets (.xls)", leave=False):
            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    if cell.ctype == xlrd.XL_CELL_FORMULA:
                        if normalized_keyword in cell.formula.lower():
                            if not copied:
                                shutil.copy(file_path, os.path.join(target_folder, os.path.basename(file_path)))
                                copied = True
                            with open(os.path.join(target_folder, 'formules_trouvees.txt'), 'a', encoding='utf-8') as f:
                                f.write(f"Files: {os.path.basename(file_path)}\nSheet: {sheet.name}\nCell: {xlrd.formula.cellname(row_idx, col_idx)}\nFormula: {cell.formula}\n\n")
                            raise StopIteration
    except StopIteration:
        pass
    except Exception as e:
        print(f"❌ Error while processing .xls formulas in {file_path} : {e}")
    return copied


def analyze_formulas_xlsm(file_path: str, keyword: str, target_folder: str) -> bool:
    """
    Analyzes formulas in a .xlsm file to search for the keyword.
    """
    copied = False
    normalized_keyword = keyword.lower().lstrip("=")
    try:
        wb = load_workbook(file_path, read_only=True, data_only=False, keep_vba=True)
        for sheet in tqdm(wb.worksheets, desc=f"[{os.path.basename(file_path)}] Sheet (.xlsm)", leave=False):
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and normalized_keyword in cell.value.lower():
                        if not copied:
                            shutil.copy(file_path, os.path.join(target_folder, os.path.basename(file_path)))
                            copied = True
                        with open(os.path.join(target_folder, 'formules_trouvees.txt'), 'a', encoding='utf-8') as f:
                            f.write(f"Files: {os.path.basename(file_path)}\nSheet: {sheet.title}\nCell: {cell.coordinate}\nFormula: {cell.value}\n\n")
                        raise StopIteration
        wb.close()
    except StopIteration:
        pass
    except Exception as e:
        print(f"❌ Error while processing .xlsm formulas in {file_path} : {e}")
    return copied


def extract_macros(file_path: str, keyword: str, target_folder: str) -> bool:
    """
    Extracts VBA macros and checks if the keyword is present.
    """
    copied = False
    try:
        vbaparser = VBA_Parser(file_path)
        if vbaparser.detect_vba_macros():
            for (_, _, vba_filename, vba_code) in vbaparser.extract_macros():
                macro_name = f"{os.path.splitext(os.path.basename(file_path))[0]}_{vba_filename}.txt"
                macro_path = os.path.join(MACRO_FOLDER, macro_name)

                with open(macro_path, 'w', encoding='utf-8') as f:
                    f.write(vba_code)

                if keyword.lower() in vba_code.lower():
                    shutil.copy(macro_path, os.path.join(target_folder, macro_name))
                    copied = True
    except Exception as e:
        print(f"Error while processing macros in {file_path} : {e}")
    return copied


def process_excel_file(file_path: str, keyword: str, target_folder: str):
    """
    Traite un fichier Excel donné : macros et formules.
    """
    ext = os.path.splitext(file_path)[1].lower()
    try:
        copied = extract_macros(file_path, keyword, target_folder)

        if ext == '.xlsm':
            copied |= analyze_formulas_xlsm(file_path, keyword, target_folder)
        elif ext == '.xls':
            copied |= analyze_formulas_xls(file_path, keyword, target_folder)

        gc.collect()
    except Exception as e:
        print(f"❌ Erreur dans {file_path} : {e}")


def convert_xls_to_xlsx(input_path: str, output_folder: str) -> str:
    """
    Convertit un fichier .xls en .xlsx via LibreOffice (headless).
    """
    if not os.path.exists(input_path):
        print(f"Fichier introuvable : {input_path}")
        return None

    try:
        subprocess.run([
            'libreoffice',
            '--headless',
            '--convert-to', 'xlsx',
            '--outdir', output_folder,
            input_path
        ], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        base = os.path.splitext(os.path.basename(input_path))[0]
        output_path = os.path.join(output_folder, f"{base}.xlsx")

        if os.path.exists(output_path):
            print(f"✅ Converti : {input_path} → {output_path}")
            return output_path
    except Exception as e:
        print(f"Erreur conversion LibreOffice : {e}")
    return None


if __name__ == '__main__':
    multiprocessing.set_start_method('spawn')

    parser = argparse.ArgumentParser(description='Keyword search in Excel macros and formulas (.xls/.xlsm)')

    parser.add_argument(
        '-k', '--keyword',
        required=True,
        type=str,
        help='Keyword to search for (e.g., sum or =sum)'
    )

    parser.add_argument(
        '-s', '--source',
        default='./assets/excel',
        help='Directory containing Excel files to scan (default: ./assets/excel)'
    )

    parser.add_argument(
        '-o', '--output',
        default='./assets/macro_trouves',
        help='Directory to store results (default: ./assets/macro_trouves)'
    )
    args = parser.parse_args()

    process_all_excels_in_parallel(args.keyword, args.source, args.output)
